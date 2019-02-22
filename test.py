#!/usr/bin/python3
#coding=utf-8

import openpyxl
import random
import time
import collections
import bisect
import math
import os
import sys

import functools
import logging
import tempfile

import copy

def read_login_sheet(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb["数据列表"]

    title = []
    for i in sheet["1"]:
        print(i.value, end=" ")
        title.append(i.value)
    print("\n")
    print(title)

    name = []
    for i in sheet["2"]:
        print(i.value, end=" ")
        name.append(i.value)
    print("\n")
    print(name)

    z = zip(title, name)
    for i,j in z:
        print(i,j)

def get_random_str(len):
    return random.sample(
        'abcdefghijklmnopqrstuvwxyz0123456789', len)

def get_random_pwd(len):
    return random.sample(
        'abcdefghijklmnopqrstuvwxyz0123456789!@#$%^&*()', len)

def get_random_int(len):
    return random.sample('0123456789', len)

def get_random_mail():
    box = get_random_str(random.randint(4, 10))
    domain1 = get_random_str(random.randint(4, 10))
    domain2 = get_random_str(random.randint(4, 10))
    return box+'@'+domain1+'.'+domain2

def get_random_address():
    '''
    :return:返回随机产生的地址码字符串，GB/T2260
    '''
    # 地区
    region = (
        # 华北地区
        11,  # 北京市
        12,  # 天津市
        13,  # 河北省
        14,  # 山西省
        15,  # 内蒙古自治区
        # 东北地区：
        21,  # 辽宁省
        22,  # 吉林省
        23,  # 黑龙江省
        # 华东地区：
        31,  # 上海市
        32,  # 江苏省
        33,  # 浙江省
        34,  # 安徽省
        35,  # 福建省
        36,  # 江西省
        37,  # 山东省
        71,  # 台湾省(886)
        # 华中地区：
        41,  # 河南省
        42,  # 湖北省
        43,  # 湖南省
        # 华南地区：
        44,  # 广东省
        45,  # 广西壮族自治区
        46,  # 海南省
        81,  # 香港特别行政区（852)
        82,  # 澳门特别行政区（853)
        # 西南地区：
        51,  # 四川省
        52,  # 贵州省
        53,  # 云南省
        54,  # 西藏自治区
        50,  # 重庆市
        # 西北地区：
        61,  # 陕西省
        62,  # 甘肃省
        63,  # 青海省
        64,  # 宁夏回族自治区
        65,  # 新疆维吾尔自治区
    )
    reg = random.randint(1,len(region))

    # 市（地级市、自治州、地区、盟及直辖市所属区和县的汇总码）
    # 其中，01-20，51-70表示地级市；21-50表示地区（自治州、盟）。
    province = random.randint(1,70)

    # 县（区、县级市、旗）。
    # 01-18表示地级市、自治州、地区、盟辖县级市；
    # 21-80表示县（旗）；81-99表示省直辖县级行政单位。
    county = random.randint(1,80)

    return "{:02d}{:02d}{:02d}".format(
        reg, province, county)

def get_random_birthday():
    '''
    :return:返回随机生成的生日字符串，GB/T7408
    '''
    y = random.randint(1900, 2019)
    m = random.randint(1, 12)
    d = random.randint(1, 28)
    return "{:04d}{:02d}{:02d}".format(y,m,d)

def get_iso_7064_checksum(idcard):
    # 校验码
    checksum = (1,0,-1,9,8,7,6,5,4,3,2)

    # 加权因子
    weight_coef = (7,9,10,5,8,4,2,1,6,3,7,9,10,5,8,4,2)

    if len(idcard) < 17:
        return None
    w = 0
    for i in range(17):
        w += int(idcard[i]) * weight_coef[i]
    r = (12 - w % 11) % 11
    if r == 10:
        return idcard[:17]+'x'
    else:
        return idcard[:17] + str(r)

def get_random_idcardnumber():
    '''
    :return:返回随机产生的身份证号码，18位
    GB11643-1999《公民身份号码》
    六位数字地址码，八位数字出生日期码，三位数字顺序码和一位数字校验码
    其中：
    1）地址码，GB/T2260
    2）出生日期码，GB/T7408
    3）顺序码，第十五位到十七位，最后一位男偶女奇
    4）校验码，ISO 7064:1983.MOD 11-2计算校验码
    '''
    s = "{:s}{:s}{:02d}{:01d}".format(
        get_random_address(),
        get_random_birthday(),
        random.randint(1,99),
        random.randint(0,9))
    s = get_iso_7064_checksum(s)
    return s


def create_login_sheet(file, num):
    '''
    创建指定行数的"数据列表"
    :param file: 文件名
    :param num: 行数
    :return:
    '''
    t0 = time.time()
    wb = openpyxl.Workbook()
    ws_login = wb.create_sheet("数据列表", 0)

    title = ['userNO', 'name', 'id',
             'idCardNumber', 'description',
             'weight#int', 'gender', 'secLevel#int',
             'userOrg', 'userOrgname', 'telephone',
             'mobilePhone', 'fax', 'homeTelephoneNumber',
             'pagerTelephoneNumber', 'address',
             'postalCode', 'countryName', 'friendlyCountryName',
             'stateOrProvinceName', 'localityName',
             'streetAddress', 'email']

    name = [ '编号', '姓名', '登录名', '身份证号', '描述',
             '排序', '性别', '密级', '组织标识', '组织名称',
             '工作电话', '移动电话', '传真', '家庭电话',
             '寻呼机号码', '邮寄地址', '邮编', '国家英文缩写',
             '国家全称', '省', '市', '街道地址', '电子邮件'
             ]

    # 表头
    idx = range(1, len(title)+1)
    z = zip(title, idx)
    for t,i in z:
        ws_login.cell(row=1, column=i, value=t)

    idx = range(1, len(name)+1)
    z = zip(name, idx)
    for t,i in z:
        ws_login.cell(row=2, column=i, value=t)


    # 数据
    for i in range(2, num+2):
        # userNO 编号
        ws_login.cell(row=i, column=1,
                      value='lruser%04d' % (i-1))

        # name 姓名
        ws_login.cell(row=i, column=2,
                      value='lruser%04d' % (i-1))

        # id 登录名
        ws_login.cell(row=i, column=3, value="".join(
            random.sample('abcdefghijklmnopqrstuvwxyz!@#$%^&*()',
                          6)))

        # idCardNumber 身份证号
        ws_login.cell(row=i, column=4,
                      value=get_random_idcardnumber())

        # description 描述
        ws_login.cell(row=i, column=5,
                      value="".join(random.sample(
                          '0123456789x',10)))

        # weight#int 排序
        ws_login.cell(row=i, column=6,
                      value=random.randint(1,10))

        # gender 性别
        ws_login.cell(row=i, column=7,
                      value="".join(random.sample(
                          'MF',1)))

        # secLevel#int 密级
        ws_login.cell(row=i, column=8,
                      value=random.randint(1,10))

        # userOrg 组织标识
        ws_login.cell(row=i, column=9,
                      value="".join(random.sample(
                          '0123456789',10)))

        # userOrgname 组织名称
        ws_login.cell(row=i, column=10,
                      value="".join(random.sample(
            'abcdefghijklmnopqrstuvwxyz',10)))

        # telephone 工作电话
        ws_login.cell(row=i, column=11,
                      value="".join(random.sample(
                          '0123456789',8)))

        # mobilePhone 移动电话
        ws_login.cell(row=i, column=12,
                      value="".join(random.sample(
                          '0123456789',10)))

        # fax 传真
        ws_login.cell(row=i, column=13,
                      value="".join(random.sample(
                          '0123456789',8)))

        # homeTelephoneNumber 家庭电话
        ws_login.cell(row=i, column=14,
                      value="".join(random.sample(
                          '0123456789',8)))

        # pagerTelephoneNumber 寻呼机号码
        ws_login.cell(row=i, column=15,
                      value="".join(random.sample(
                          '0123456789',10)))

        # address 邮寄地址
        ws_login.cell(row=i, column=16,
                      value="".join(random.sample(
            'abcdefghijklmnopqrstuvwxyz',20)))

        # postalCode 邮编
        ws_login.cell(row=i, column=17,
                      value="".join(random.sample(
                          '0123456789',6)))

        # countryName 国家英文缩写
        ws_login.cell(row=i, column=18,
                      value="".join(random.sample(
                          'abcdefghijklmnopqrstuvwxyz',
                          3)))

        # friendlyCountryName 国家全称
        ws_login.cell(row=i, column=19,
                      value="".join(random.sample(
                          'abcdefghijklmnopqrstuvwxyz',
                          20)))

        # stateOrProvinceName 省
        ws_login.cell(row=i, column=20,
                      value="".join(random.sample(
                          'abcdefghijklmnopqrstuvwxyz',
                          10)))

        # localityName 市
        ws_login.cell(row=i, column=21,
                      value="".join(random.sample(
                          'abcdefghijklmnopqrstuvwxyz',
                          10)))

        # streetAddress 街道地址
        ws_login.cell(row=i, column=22,
                      value="".join(random.sample(
                          'abcdefghijklmnopqrstuvwxyz',
                          20)))

        # email 电子邮件
        ws_login.cell(row=i, column=23,
                      value="".join(random.sample(
                          'abcdefghijklmnopqrstuvwxyz',
                          5))+ r"@163.com")

    try:
        wb.save(file)
    except PermissionError as e:
        print(e)
    finally:
        t1 = time.time()
        print("耗时%0.2f秒." % (t1 - t0))




def str_format():
    title = ("剧名", "剩余票数", "票价", "日期")
    rec = (("凤还巢", 10, 30.16, "2019-2-5 19:30:00"),
           ("七星灯", 104, 40.52, "2019-2-7 19:30:00"),
           ("莱茵的黄金", 255, 30.16, "2019-2-6 18:00:00"),
           ("乌盆记", 3, 66.16, "2019-2-10 19:30:00"))
    sche = []
    for i in range(4):
        sche.append(zip(title, rec[i]))

    print("{:20s} {:10s} {:15s} {:20s}".format(title[0],
                                               title[1],
                                               title[2],
                                               title[3]))

    print("-"*60)
    for i in range(len(rec)):
        print("{:<20s} {:>10d} {:>10.2f} {:<20s}".format(rec[i][0],
                                                 rec[i][1],
                                                 rec[i][2],
                                                 rec[i][3]))
    

def iter_test():
    L = list(tuple("ABCDEF"))
    print(L)
    L[2:5] = ["x", "y"]
    print(L)

    L = list(tuple("ABCDEF"))
    L[2:4] = []
    print(L)

    L = list(tuple("ABCDEF"))
    del L[2:4]
    print(L)


def grade(score, breakpoints=[60, 70, 80, 90], grades='FDCBA'):
    # i = bisect.bisect(breakpoints, score)
    i = bisect.bisect_left(breakpoints, score)

    return grades[i]


def fun_add(info):
    print("fun_add() %s" % info)

def fun_modify(info):
    print("fun_modify() %s" % info)

def fun_del(info):
    print("fun_del() %s" % info)

def fun1(fun, param):
    functions = dict(a=fun_add, m=fun_modify, d=fun_del)
    functions[fun](param)


def iter_fun1(d):
    for k in sorted(d):
        yield k, d[k]

def iter_fun2(d):
    return (((k, d[k]) for k in sorted(d)))


def load_modules(targe_path=''):
    '''
    【动态导入模块】
    对程序所在路径所有文件进行迭代；
    每个以.py为扩展名，且名称中包含'magic'的文件，获取模块名；
    若模块名是有效标识符，就说明其是一个可用的模块名；
    若模块名尚未存于sys.modules字典中，则尝试将其导入。
    :param targe_path:
    :return:
    '''
    modules=[]
    for name in os.listdir(os.path.dirname(__file__) or '.'):
        if name.endswith(".py") and "magic" in name.lower():
            filename = name
            name = os.path.splitext(name)[0] # 模块名
            if name.isidentifier() and name not in sys.modules:
                fh = None
                try:
                    fh = open(filename, "r", encoding='utf-8')
                    code = fh.read()
                    module = type(sys)(name) # 创建一个新模块
                    sys.modules[name] = module # 将模块添加到字典中
                    exec(code, module.__dict__)
                    modules.append(module)
                except (EnvironmentError, SyntaxError) as err:
                    sys.modules.pop(name, None)
                    print(err)
                finally:
                    if fh is not None:
                        fh.close()
    return modules

def get_function(module, function_name):
    '''
    对模块对象调用getattr()，若不存在所需函数，就产生AttributeError异常；
    若存在所需函数，就调用hasattr()检查函数是否具备__call__属性，可调用对象都具备__call__属性；
    若函数可调用，就返回给调用者。
    :param module:
    :param function_name:
    :return:
    '''
    function = get_function.cache.get(
        (module, function_name), None)
    if function is None:
        try:
            function = getattr(module, function_name)
            if not hasattr(function, "__call__"):
                # 可调用对象都具备__call__属性
                raise AttributeError()
            get_function.cache[module, function_name] = function
        except AttributeError:
            function = None
    return function
get_function.cache={}






class A():
    d1 = None
    __d2 = None  #共享的类变量

    def __init__(self, d1, d2):
        self.d1 = d1
        self.__d2 = d2

    def fun(self, s): #类方法
        print("fun() %s, %s, %s" %
              (s, str(self.d1), str(self.__d2)))

    @property
    def d2(self):
        return self.__d2

    @d2.setter
    def d2(self, d):
        self.__d2 = d

class B():
    pass

class C(A,B):
    def fun(self,s):
        A.fun(self,s)
        print("C.fun() %s" % s)

class D():
    __slots__ = ("x", "y")
    pass

class Point:
    __slots__ = ("x", "y")
    def __init__(self, x=0, y=0):
        self.x = x
        self.y = y

class Const:
    def __setattr__(self, key, value):
        if key in self.__dict__:
            raise ValueError("cannot change a const attribute")
        self.__dict__[key] = value

    def __delattr__(self, item):
        if item in self.__dict__:
            raise ValueError("cannot delete a const attribute")
        raise AttributeError("'{0}' object has no attribute '{1}'"
                             .format(self.__class__.__name__, item))

class Strip:
    def __init__(self, characters):
        self.characters = characters

    def __call__(self, string):
        return string.strip(self.characters)

def make_strip_function(charcters):
    def strip_function(string):
        return string.strip(charcters)
    return strip_function


def tag(tag_name):
    def add_tag(content):
        return "<{0}>{1}</{0}>".format(tag_name, content)
    return add_tag



def makebold(fn):
    def wrapped():
        return "<b>" + fn() + "</b>"
    return wrapped

def makeitalic(fn):
    def wrapped():
        return "<i>" + fn() + "</i>"
    return wrapped

@makebold
@makeitalic
def hello():
    return "hello world"

def timeit(fn):
    def wrapper():
        start = time.perf_counter()
        fn()
        end = time.perf_counter()
        print("Time elapsed:{:.3e}s".format(end - start))
    return wrapper

@timeit
def foo():
    print("in foo()")



#
# if __debug__: # 调试模式下，全局变量__debug__为True
#     logger = logging.getLogger("Logger")
#     logger.setLevel(logging.DEBUG)
#     handler = logging.FileHandler(os.path.join(
#         tempfile.gettempdir(), "logged.log"))
#     logger.addHandler(handler)
#
#     def logged(fn):
#         @functools.wraps(fn)
#         def wrapper(*args, **kwargs):
#             log = "called:" + fn.__name__ + "("
#             log += ",".join(["{0!r}".format(a) for a in args] +
#                             ["{0!s}={1!r}".format(k,v)
#                              for k,v in kwargs.items()])
#             result = exception = None
#             try:
#                 result = fn(*args, **kwargs)
#                 return result
#             except Exception as err:
#                 exception = err
#             finally:
#                 log += ((")->" + str(result) if exception is None else ") {0}:{1}".format(type(exception), exception)))
#                 logger.debug(log)
#                 if exception is not None:
#                     raise exception
#             return wrapper
# else:
#     def logged(fn):
#         def wrapper():
#             fn()
#         return fn

def logged(fn):
    @functools.wraps(fn)
    def wrapper(*args, **kwargs):
        log = "called:" + fn.__name__ + "("
        log += ",".join(["{0!r}".format(a) for a in args] +
                        ["{0!s}={1!r}".format(k,v)
                         for k,v in kwargs.items()])
        result = exception = None
        try:
            result = fn(*args, **kwargs)
            return result
        except Exception as err:
            exception = err
        finally:
            log += ((")->" + str(result) if exception is None else ") {0}:{1}".format(type(exception), exception)))
            logger.debug(log)
            if exception is not None:
                raise exception
    return wrapper

@logged
def fun111():
    print("fun111()")


def with_test(file_name):
    try:
        with open(file_name) as fi, \
                open(r'tmp1','w') as fo1, \
                open(r'tmp2','w') as fo2:
            for line in fi:
                print(line)
                fo1.write("fo1:"+line)
                fo2.write("fo2:"+line)
    except EnvironmentError as err:
        print(err)

def get_id(target):
    print("*"*10)
    print(target, id(target))
    # if hasattr(target, "__iter__"):
    if isinstance(target, collections.Iterable):
        for i in iter(target):
            print(i, id(i))


def fun11():
    lst1 = ['a','b','c']
    lst2 = [1,2,lst1]
    lst3 = lst2
    get_id(lst1)
    get_id(lst2)
    get_id(lst3)
    print("-"*30)

    lst3 = copy.copy(lst2)
    get_id(lst1)
    get_id(lst2)
    get_id(lst3)
    print("-"*30)

    lst3 = copy.deepcopy(lst2)
    get_id(lst1)
    get_id(lst2)
    get_id(lst3)
    print("-"*30)

class AtomicList:
    def __init__(self, alist, shallow_copy=True):
        self.original = alist
        self.shallow_copy = shallow_copy

    def __enter__(self):
        self.modified = (self.original[:] if self.shallow_copy
                         else copy.deepcopy(self.original))
        return self.modified

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self.original[:] = self.modified

def AtomicList_test():
    items = list(range(10))
    index = 12
    print(items)
    try:
        with AtomicList(items) as atomic:
            atomic.append(123456)
            del atomic[3]
            atomic[8] =8463
            atomic[index] = 41843
    except (AttributeError, IndexError, ValueError) as err:
        print("no changes applied:", err)
    print(items)

class Foo(object):
    """docstring for Foo"""

    def __init__(self, arg=''):
        super(Foo, self).__init__()
        self.arg = arg

    def foo(self):
        print(self)
        print('foo:', 123)


class Descriptor(object):
    def __init__(self, value):
        self.value = value

    def __get__(self, instance, owner):
        print("访问属性")
        return self.value

    def __set__(self, instance, value):
        print("设置属性值")
        self.value = value

class Myclass(object):
    desc = Descriptor(5)



if __name__=="__main__":
    # str_format()
    # a = 10
    # b = 20
    # print(a and b)
    # print(a or b)
    # print(not a)

    # 用户登录信息
    # create_login_sheet(r"D:\wangbin\my_workspace\python_intro\login.xlsx", 20000)
    # read_login_sheet(r"./user.xlsx")

    # SJZTPZ
    # 设备、电连接器、电缆

    # l =  [grade(score) for score in [33, 99, 77, 70, 89, 90, 100]]
    # print(l)

    # b = B()
    # print(dir(b))
    #
    # d = D()
    # print(dir(d))
    #
    # point=Point()
    # print(dir(point))

    # a = A("aa", 12)
    # a.fun("test")
    # a._A__d2 = 'AA'
    # a.fun("test1")
    # a.d2 = 'ss'  #写属性
    # print(a.d2)  #读属性
    # a.fun("test2")
    # a.d1 = 100
    # a.fun("test3")
    # print(dir(A))

    # for i in range(10):
    #     fun1(random.choice(('a', 'm', 'd')),
    #          str(random.randrange(0,100)))

    # d = dict(a="add", m="modify", d="del")
    # it = iter_fun1(d)
    # print(next(it))
    # print(next(it))
    # print(next(it))
    #
    # for i in list(iter_fun1(d)):
    #     print(i)
    #
    # for i in tuple(iter_fun1(d)):
    #     print(i)
    # modules = load_modules(r'D:\wangbin\my_workspace\python_intro')
    # if modules is not None:
    #     for i in modules:
    #         print(i)
    #     fun = get_function(modules[0], "get_random_address")
    #     if fun is not None:
    #         print("{}".format(fun()))
    # print("over")

    # print(hello())
    # foo()

    # logger = logging.getLogger("Logger")
    # logger.setLevel(logging.DEBUG)
    # handler = logging.FileHandler(os.path.join(
    #     tempfile.gettempdir(), "logged.log"))
    # logger.addHandler(handler)
    #
    # fun111()

    # c = Const()
    # c.name = 'test'
    # print(c.name)
    # c.name = 'aaa'  # 不能修改只读属性
    # del c.name #不能删除只读属性
    #
    # strip_punctuation = Strip(",;:.!?")
    # print(strip_punctuation("Land ahoy!!.;"))

    # strip_punctuation = make_strip_function(",;:.!?")
    # print(strip_punctuation("Land ahoy!!.;"))

    # add_html = tag("html")
    # add_head = tag("head")
    # add_title = tag("title")
    # add_body = tag("body")
    # ctx = add_html(add_head(add_title("welcome"))+add_body("嘻嘻哈哈"))
    # print(ctx)

    # with_test(r"D:\wangbin\my_workspace\python_intro\jmeter.log")
    # fun11()
    # AtomicList_test()

    # print(Foo.foo)
    #
    # print(Foo().foo)
    #
    # print(Foo.foo.__get__)

    #访问类属性
    # print(Myclass.desc)
    # print("*"*10)
    # Myclass.desc = 6
    # print(Myclass.desc)
    # print("*"*10)

    #访问实例属性
    for i in range(10):
        print(Myclass.desc)
        myClass = Myclass()
        print(myClass.desc)
        myClass.desc = i
        print(myClass.desc)
        print('*'*10)




